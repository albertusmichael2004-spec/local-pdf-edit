from __future__ import annotations

from pathlib import Path

import fitz

from backend.core.errors import ConversionError
from backend.core.progress import report_fraction, report_progress

def clean_excel_value(value: object) -> object:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    # XLSX does not permit most ASCII control characters.
    return ''.join(ch for ch in text if ch in '\t\n\r' or ord(ch) >= 32)

def pdf_to_xlsx(input_path: Path, output_path: Path) -> tuple[int, int]:
    """Best-effort table extraction to XLSX with a resilient text fallback."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ConversionError("openpyxl is not installed. Run pip install -r requirements.txt.") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    table_count = 0
    page_count = 0
    try:
        plumber_error = None
        try:
            import pdfplumber
            with pdfplumber.open(str(input_path)) as pdf:
                page_count = len(pdf.pages)
                for page_idx, page in enumerate(pdf.pages, start=1):
                    ws = workbook.create_sheet(title=f"Page {page_idx}"[:31])
                    row_cursor = 1
                    try:
                        tables = page.extract_tables() or []
                    except Exception as exc:
                        tables = []
                        plumber_error = exc
                    if tables:
                        for t_idx, table in enumerate(tables, start=1):
                            table_count += 1
                            if len(tables) > 1:
                                ws.cell(row=row_cursor, column=1, value=f"Table {t_idx}")
                                row_cursor += 1
                            for row in table:
                                for col_idx, value in enumerate(row or [], start=1):
                                    ws.cell(row=row_cursor, column=col_idx, value=clean_excel_value(value))
                                row_cursor += 1
                            row_cursor += 2
                    else:
                        text = ""
                        try:
                            text = page.extract_text() or ""
                        except Exception as exc:
                            plumber_error = exc
                        for line in text.splitlines():
                            ws.cell(row=row_cursor, column=1, value=clean_excel_value(line))
                            row_cursor += 1
                    ws.freeze_panes = "A1"
                    report_fraction("Extracting PDF tables", page_idx, page_count, 24, 88)
        except Exception as exc:
            plumber_error = exc

        # If pdfplumber failed completely or created no pages, use PyMuPDF so
        # the conversion never silently becomes a network-looking "failed to fetch".
        if not workbook.sheetnames:
            with fitz.open(input_path) as doc:
                if doc.needs_pass:
                    raise ConversionError("Encrypted PDF. Unlock it before conversion.")
                page_count = doc.page_count
                for page_idx, page in enumerate(doc, start=1):
                    ws = workbook.create_sheet(title=f"Page {page_idx}"[:31])
                    text = page.get_text("text") or ""
                    for row_idx, line in enumerate(text.splitlines() or [""], start=1):
                        ws.cell(row=row_idx, column=1, value=clean_excel_value(line))
                    ws.freeze_panes = "A1"
                    report_fraction("Extracting PDF text", page_idx, page_count, 24, 88)

        if not workbook.sheetnames:
            ws = workbook.create_sheet(title="Page 1")
            ws["A1"] = "No extractable content found."
            page_count = 1
        report_progress("Saving Excel workbook", percent=94)
        workbook.save(output_path)
        return len(workbook.sheetnames), table_count
    except ConversionError:
        raise
    except Exception as exc:
        raise ConversionError(f"PDF to Excel conversion failed: {exc}") from exc
    finally:
        try:
            workbook.close()
        except Exception:
            pass
