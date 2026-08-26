from __future__ import annotations

from pathlib import Path

from backend.core.errors import ConversionError
from backend.services.shared.renderers.reportlab_common import reportlab_imports, safe_para_text

def render_xlsx_to_pdf(input_path: Path, output_path: Path) -> str:
    if input_path.suffix.lower() != ".xlsx":
        raise ConversionError("Legacy .xls files require LibreOffice. Save the file as .xlsx or install LibreOffice.")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ConversionError("openpyxl is not installed. Run pip install -r requirements.txt.") from exc
    rl = reportlab_imports()
    try:
        wb = load_workbook(str(input_path), data_only=True, read_only=True)
        styles = rl["getSampleStyleSheet"]()
        story = []
        for sheet_index, ws in enumerate(wb.worksheets):
            story.append(rl["Paragraph"](safe_para_text(ws.title), styles["Heading2"]))
            rows = []
            max_cols = min(ws.max_column or 1, 20)
            max_rows = min(ws.max_row or 1, 5000)
            for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                rows.append([safe_para_text(v) for v in row])
            while rows and not any(rows[-1]):
                rows.pop()
            if rows:
                table = rl["Table"](rows, repeatRows=1, hAlign="LEFT")
                table.setStyle(rl["TableStyle"]([
                    ("GRID", (0,0), (-1,-1), 0.25, rl["colors"].lightgrey),
                    ("FONTSIZE", (0,0), (-1,-1), 6.5),
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                    ("BACKGROUND", (0,0), (-1,0), rl["colors"].whitesmoke),
                ]))
                story.append(table)
            else:
                story.append(rl["Paragraph"]("(Empty sheet)", styles["BodyText"]))
            if sheet_index < len(wb.worksheets) - 1:
                story.append(rl["PageBreak"]())
        wb.close()
        pdf = rl["SimpleDocTemplate"](str(output_path), pagesize=rl["landscape"](rl["A4"]), rightMargin=24, leftMargin=24, topMargin=28, bottomMargin=28)
        pdf.build(story)
        return "Built-in XLSX renderer"
    except ConversionError:
        raise
    except Exception as exc:
        raise ConversionError(f"Built-in XLSX conversion failed: {exc}") from exc
